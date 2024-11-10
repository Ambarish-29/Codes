import openpyxl as xl
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import Protection
import win32api
import win32con
import win32com.client as win32
import os
import time

# Function to protect sheet
def protect_sheet(wb, ws):
    ws.protection = SheetProtection(selectLockedCells=True, selectUnlockedCells=True)
    ws.protection.enable()

    # Lock all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    file_path = os.path.abspath('hidden_ribbon_workbook_macfin.xlsx')
    wb.save(file_path)
    print(f"Protected Excel file created at: {file_path}")
    return file_path

# Function to inject VBA to prevent printing
def prevent_printing(wb_with_macro):
    vba_code = """
        Private Sub Workbook_BeforePrint(Cancel As Boolean)
            Cancel = True
            MsgBox "Printing is disabled for this workbook."
        End Sub
    """
    vba_module = wb_with_macro.VBProject.VBComponents("ThisWorkbook").CodeModule
    vba_module.DeleteLines(1, vba_module.CountOfLines)  # Clear existing code
    vba_module.AddFromString(vba_code)

# Function to inject VBA to prevent saving after the first save
def prevent_saving(wb_with_macro):
    vba_code = """
    Private Sub Workbook_BeforeSave(ByVal SaveAsUI As Boolean, Cancel As Boolean)
        If ThisWorkbook.Saved = False Then
            ' Allow initial save
            ThisWorkbook.Saved = True  ' Mark as saved
        Else
            ' Prevent further saves
            Cancel = True
        End If
    End Sub
    """
    vba_module = wb_with_macro.VBProject.VBComponents("ThisWorkbook").CodeModule
    vba_module.AddFromString(vba_code)

# Main function
def main():
    # Step 1: Create a new Excel workbook
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "ProtectedSheet"
    ws['A1'] = "This content is protected, printing is disabled, and saving is restricted."

    # Step 2: Protect the sheet
    file_path = protect_sheet(wb, ws)

    # Step 3: Enable VBA project access in the registry
    excel_version = "16.0"
    registry_path = fr"Software\Microsoft\Office\{excel_version}\Excel\Security"
    try:
        key = win32api.RegOpenKeyEx(win32con.HKEY_CURRENT_USER, registry_path, 0, win32con.KEY_ALL_ACCESS)
        win32api.RegSetValueEx(key, "AccessVBOM", 0, win32con.REG_DWORD, 1)
        win32api.RegCloseKey(key)
        print(f"Successfully enabled programmatic access to VBA for Excel version {excel_version}.")
    except Exception as e:
        print(f"Error enabling programmatic access: {e}")
        return

    # Step 4: Open the workbook with win32com to inject VBA
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    wb_with_macro = excel.Workbooks.Open(file_path)
    time.sleep(5)  # Delay to ensure the workbook is fully initialized

    # Step 5: Inject VBA to prevent printing
    prevent_printing(wb_with_macro)

    # Step 6: Inject VBA to prevent saving after the first save
    prevent_saving(wb_with_macro)

    # Step 7: Save the workbook as a macro-enabled workbook (.xlsm)
    macro_enabled_path = os.path.abspath('hidden_ribbon_workbook_macros_combined.xlsm')
    wb_with_macro.SaveAs(macro_enabled_path, FileFormat=52)  # 52 is for .xlsm format
    time.sleep(3)  # Add a delay to ensure saving is complete

    # Step 8: Close the workbook and quit Excel
    wb_with_macro.Close(SaveChanges=True)
    excel.Quit()

    print(f"Macro-enabled workbook created at: {macro_enabled_path}")

# Run the main function
main()
